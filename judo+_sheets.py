from openpyxl import load_workbook
from config import Config
import yaml
import os

# =========================================================
# CONFIG
# =========================================================

os.makedirs(Config.excel_output_folder(), exist_ok=True)


# =====================================================
# LOAD YAML
# =====================================================

def load_yaml(file_path, default=None):
    """
    Generic YAML loader.

    :param file_path: Path to the YAML file
    :type file_path: str
    :param default: Default value if file is missing or empty
    :type default: any
    :return: Parsed YAML content or default value
    :rtype: dict | list
    """
    if not os.path.exists(file_path):
        print(f"Warning: file not found -> {file_path}")
        return default if default is not None else {}

    with open(file_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    if data is None:
        return default if default is not None else {}

    return data


# =====================================================
# BUILD GROUP LOOKUP FROM GROUPED YAML
# =====================================================

def build_group_lookup(grouped_data, sex_label):
    """
    Build a lookup dictionary for fast group → athlete resolution.

    Key format:
        (Sex, AgeTier, WeightClass, GroupName)

    Example:
        ("Male", "Benjamim", "-38", "Group1")

    :param grouped_data: Parsed grouped YAML structure
    :type grouped_data: dict
    :param sex_label: Gender label ("Male" / "Female")
    :type sex_label: str
    :return: Lookup dictionary mapping group identifiers to athlete lists
    :rtype: dict
    """
    lookup = {}

    if not isinstance(grouped_data, dict):
        return lookup

    for age_tier, weight_classes in grouped_data.items():
        if not isinstance(weight_classes, dict):
            continue

        for weight_class, groups_dict in weight_classes.items():
            if not isinstance(groups_dict, dict):
                continue

            for group_name, athletes in groups_dict.items():
                if not str(group_name).startswith("Group"):
                    continue

                if not isinstance(athletes, list):
                    athletes = []

                key = (sex_label, age_tier, weight_class, group_name)
                lookup[key] = athletes

    return lookup


def build_combined_group_lookup():
    """
    Build a combined lookup from both male and female YAML files.

    :return: Combined lookup dictionary (male + female groups)
    :rtype: dict
    """
    male_data = load_yaml(Config.male_grouped_file(), default={})
    female_data = load_yaml(Config.female_grouped_file(), default={})

    lookup = {}
    lookup.update(build_group_lookup(male_data, Config.MALE))
    lookup.update(build_group_lookup(female_data, Config.FEMALE))

    return lookup


# =====================================================
# POPULATE EXCEL
# =====================================================

def populate_group_table(group_data,
                         athletes,
                         template_file_2x,
                         template_file_5,
                         template_file_2,
                         output_file_pattern
):
    """
    Populate an Excel sheet using the appropriate template for the group size.

    Template selection:
    - 2 athletes → BO3 template
    - 3–5 athletes → standard pool template
    - >5 athletes → 2x template (split layout)

    :param group_data: Metadata for the group (mat, tier, weight, etc.)
    :type group_data: dict
    :param athletes: List of athletes in the group
    :type athletes: list[dict]
    :param template_file_2x: Path to large-group template
    :type template_file_2x: str
    :param template_file_5: Path to 3–5 athlete template
    :type template_file_5: str
    :param template_file_2: Path to BO3 template (2 athletes)
    :type template_file_2: str
    :param output_file_pattern: Filename pattern for output Excel files
    :type output_file_pattern: str
    :return: None
    :rtype: None
    """
    group_size = group_data["GroupSize"]
    mat_number = group_data["MatNumber"]
    age_tier_full = group_data["AgeTier"]
    age_tier_short = group_data["AgeTierShort"].replace("/", "-")
    weight_class = group_data["WeightClass"]

    group_short = group_data.get(
        "GroupShort", group_data["GroupName"].replace("Group", "G")
    )

    # =====================================================
    # CASE 1: BO3 (2 athletes)
    # =====================================================
    if group_size == 2:
        workbook = load_workbook(template_file_2)
        sheet = workbook.active

        sheet['B3'] = Config.LOC_DATE
        sheet['F13'] = group_short.replace("G", "")
        sheet['G4'] = age_tier_full
        sheet['L4'] = str(weight_class) + " kg"
        sheet['L2'] = mat_number

        for i, athlete in enumerate(athletes, start=8):
            sheet[f'C{i}'] = athlete.get('Name', '')
            sheet[f'D{i}'] = str(athlete.get('Weight', '')) + " kg"
            sheet[f'E{i}'] = athlete.get('Club', '')
    
    # =====================================================
    # CASE 2: STANDARD GROUP (3–5 athletes)
    # =====================================================
    elif 3 <= group_size <= 5:
        workbook = load_workbook(template_file_5)
        sheet = workbook.active

        sheet['C3'] = Config.LOC_DATE
        sheet['M17'] = group_short.replace("G", "")
        sheet['K4'] = age_tier_full
        sheet['Q4'] = str(weight_class) + " kg"
        sheet['Q2'] = mat_number

        for i, athlete in enumerate(athletes, start=9):
            sheet[f'B{i}'] = athlete.get('Name', '')
            sheet[f'C{i}'] = str(athlete.get('Weight', '')) + " kg"
            sheet[f'D{i}'] = athlete.get('Club', '')

    # =====================================================
    # CASE 2: STANDARD GROUP (3–5 athletes) --- UNUSED ---
    # =====================================================
    elif 3 <= group_size <= 5:
        workbook = load_workbook(template_file_5)
        sheet = workbook.active

        sheet['C3'] = Config.LOC_DATE
        sheet['M17'] = group_short.replace("G", "")
        sheet['K4'] = age_tier_full
        sheet['Q4'] = str(weight_class) + " kg"
        sheet['Q2'] = mat_number

        for i, athlete in enumerate(athletes, start=9):
            sheet[f'B{i}'] = athlete.get('Name', '')
            sheet[f'C{i}'] = str(athlete.get('Weight', '')) + " kg"
            sheet[f'D{i}'] = athlete.get('Club', '')

    output_filename = os.path.join(
        Config.excel_output_folder(),
        output_file_pattern.format(
            age_tier_short=age_tier_short,
            sex=group_data["Sex"][0],
            group_number=group_short,
            weight_class=weight_class
        )
    )

    workbook.save(output_filename)
    print(f"{group_short} on mat {mat_number} saved to {output_filename}")


# =====================================================
# MAIN
# =====================================================

def generate_excel_from_yaml(yaml_file_path):
    """
    Generate Excel fight sheets from mat distribution YAML.

    Pipeline:
    1. Load mat distribution YAML
    2. Load grouped athlete YAMLs (male + female)
    3. Build lookup table for fast access
    4. Match assignments → athletes
    5. Populate correct Excel templates

    :param yaml_file_path: Path to mat distribution YAML
    :type yaml_file_path: str
    :return: None
    :rtype: None
    """
    mat_data = load_yaml(yaml_file_path, default={})

    assignments = mat_data.get("Assignments", [])
    if not assignments:
        print("No assignments found in mat distribution YAML.")
        return

    group_lookup = build_combined_group_lookup()

    for group in assignments:
        key = (
            group["Sex"],
            group["AgeTier"],
            group["WeightClass"],
            group["GroupName"]
        )

        athletes = group_lookup.get(key, [])

        if not athletes:
            print(f"Warning: no athletes found for key={key}")

        populate_group_table(
            group_data=group,
            athletes=athletes,
            template_file_2x=Config.TEMPLATE_2X,
            template_file_5=Config.TEMPLATE_5,
            template_file_2=Config.TEMPLATE_BO3,
            output_file_pattern=Config.EXCEL_PATTERN
        )


if __name__ == "__main__":
    generate_excel_from_yaml(Config.mat_distribution_yaml())